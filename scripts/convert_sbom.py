"""
Convert CycloneDX 1.6 JSON SBOM into an XLSX workbook.
"""

import argparse
import json
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font


def _format_tool_entry(tool: dict) -> str | None:
    if not isinstance(tool, dict):
        return None

    name = tool.get("name", "Unknown Tool")
    version = tool.get("version", "")
    group = tool.get("group", "")
    author = tool.get("author", "")
    license_ids = [
        lic.get("license", {}).get("id")
        for lic in tool.get("licenses", [])
        if lic.get("license", {}).get("id")
    ]

    parts = [name]
    if group:
        parts.append(f"(Group: {group})")
    if version:
        parts.append(f"v{version}")
    if author:
        parts.append(f"by {author}")
    if license_ids:
        parts.append(f"[License: {', '.join(license_ids)}]")

    return " ".join(parts)


def _flatten_component(component: dict, lookup: dict, parent_ref: str = "") -> None:
    bom_ref = component.get("bom-ref") or (
        f"{parent_ref}|{component.get('name', '')}@{component.get('version', '')}"
    )
    lookup[bom_ref] = component

    for nested in component.get("components", []):
        _flatten_component(nested, lookup, bom_ref)


def _build_component_row(bom_ref: str, component: dict) -> dict:
    name = component.get("name", "")
    version = component.get("version", "")
    group = component.get("group", "")
    description = component.get("description", "")
    package_name = f"{group + '/' if group else ''}{name}"

    licenses = component.get("licenses", [])
    if licenses:
        license_info = ", ".join(
            lic.get("license", {}).get("id", "")
            for lic in licenses
            if lic.get("license", {}).get("id")
        )
        license_info = license_info or "No Assertion"
    else:
        license_info = "No Assertion"

    author = component.get("author") or "Not Listed"
    purl = component.get("purl", "")

    hash_value = ""
    for ext_ref in component.get("externalReferences", []):
        if ext_ref.get("type") == "distribution":
            hashes = ext_ref.get("hashes", [])
            if hashes:
                hash_value = f"{hashes[0]['alg']}:{hashes[0]['content']}"
                break

    path_value = next(
        (
            prop["value"]
            for prop in component.get("properties", [])
            if prop.get("name", "").endswith("package:path")
        ),
        "",
    )

    return {
        "Package Name": package_name,
        "Version": version,
        "Author": author,
        "License": license_info,
        "Description": description,
        "PURL": purl,
        "Path": path_value,
        "BOM Reference": bom_ref,
        "Hash Value": hash_value,
    }


def sbom_to_excel(input_json: Path, output_excel: Path) -> None:
    with input_json.open(encoding="utf-8") as f:
        sbom = json.load(f)

    metadata = sbom.get("metadata", {})
    metadata_rows = [
        ["Timestamp", metadata.get("timestamp", "Not Available")],
        ["Component Name", metadata.get("component", {}).get("name", "Not Available")],
        [
            "Component Version",
            metadata.get("component", {}).get("version", "Not Available"),
        ],
        ["Component Type", metadata.get("component", {}).get("type", "Not Available")],
        [
            "Component BOM Ref",
            metadata.get("component", {}).get("bom-ref", "Not Available"),
        ],
    ]

    tools_section = metadata.get("tools", {})
    tool_entries = []
    for tool in tools_section.get("components", []):
        formatted = _format_tool_entry(tool)
        if formatted is not None:
            tool_entries.append(formatted)

    if tool_entries:
        metadata_rows.append(["Tools Used", tool_entries[0]])
        for tool_entry in tool_entries[1:]:
            metadata_rows.append(["", tool_entry])
    else:
        metadata_rows.append(["Tools Used", "None"])

    metadata_df = pd.DataFrame(metadata_rows)

    explanation_rows = [
        ["SBOM and Vulnerability Report Explanation"],
        [""],
        ["SBOM Generation Process:"],
        [
            "This SBOM was generated from CycloneDX JSON output and converted to an Excel workbook "
            "to support review and ATO submission workflows."
        ],
        [
            "The pipeline builds the image, scans dependencies and the container, merges SBOM sources, "
            "and exports this human-readable artifact."
        ],
        [""],
        ["Workbook Structure:"],
        ["- Metadata: root component and generation metadata."],
        ["- Components: all identified software components."],
        ["- Dependencies: dependency graph edges between components."],
    ]
    explanation_df = pd.DataFrame(explanation_rows)

    component_lookup: dict = {}
    for component in sbom.get("components", []):
        _flatten_component(component, component_lookup)

    component_rows = [
        _build_component_row(bom_ref, comp)
        for bom_ref, comp in component_lookup.items()
    ]
    components_df = pd.DataFrame(component_rows)
    if not components_df.empty:
        components_df = components_df.sort_values(by=["Package Name", "Version"])

    dependency_rows = [
        {"Component": dependency.get("ref", ""), "Depends On": dep}
        for dependency in sbom.get("dependencies", [])
        for dep in dependency.get("dependsOn", [])
    ]
    dependencies_df = pd.DataFrame(dependency_rows)

    with pd.ExcelWriter(output_excel) as writer:
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False, header=False)
        explanation_df.to_excel(
            writer, sheet_name="SBOM Explanation", index=False, header=False
        )
        components_df.to_excel(writer, sheet_name="Components", index=False)
        dependencies_df.to_excel(writer, sheet_name="Dependencies", index=False)

    workbook = openpyxl.load_workbook(output_excel)
    metadata_sheet = workbook["Metadata"]
    bold_font = Font(bold=True)
    for row in range(1, metadata_sheet.max_row + 1):
        metadata_sheet.cell(row=row, column=1).font = bold_font
    workbook.save(output_excel)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert CycloneDX 1.6 SBOM JSON file into an Excel workbook"
    )
    parser.add_argument("-i", "--input", required=True, help="Input SBOM JSON file")
    parser.add_argument(
        "-o",
        "--output",
        default="cyclonedx_report.xlsx",
        help="Output XLSX file",
    )
    args = parser.parse_args()

    input_file = Path(args.input)
    if not input_file.exists():
        raise SystemExit(f"Input file not found: {input_file}")

    sbom_to_excel(input_file, Path(args.output))
