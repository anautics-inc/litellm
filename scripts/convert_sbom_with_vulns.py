import json
import pandas as pd
from pathlib import Path
import argparse
import openpyxl
from openpyxl.styles import Font

def build_vulnerabilities_df(sbom):
    columns = [
        "ID", "Source", "Description", "Severity", "CVSS Score",
        "Affected Package", "Analysis State", "Analysis Detail", "Advisory URLs"
    ]
    vulns = sbom.get("vulnerabilities", [])
    rows = []
    for vuln in vulns:
        vid = vuln.get("id", "")
        source = vuln.get("source", {}).get("name", "")
        description = vuln.get("description", "")
        ratings = vuln.get("ratings", [])
        severity = ratings[0].get("severity", "N/A") if ratings else "N/A"
        score = ratings[0].get("score", "N/A") if ratings else "N/A"
        analysis = vuln.get("analysis", {})
        state = analysis.get("state", "")
        detail = analysis.get("detail", "")
        advisory_urls = ";".join(a.get("url", "") for a in vuln.get("advisories", []))
        affects = vuln.get("affects", [])
        if affects:
            for affect in affects:
                rows.append({
                    "ID": vid,
                    "Source": source,
                    "Description": description,
                    "Severity": severity,
                    "CVSS Score": score,
                    "Affected Package": affect.get("ref", ""),
                    "Analysis State": state,
                    "Analysis Detail": detail,
                    "Advisory URLs": advisory_urls,
                })
        else:
            rows.append({
                "ID": vid,
                "Source": source,
                "Description": description,
                "Severity": severity,
                "CVSS Score": score,
                "Affected Package": "",
                "Analysis State": state,
                "Analysis Detail": detail,
                "Advisory URLs": advisory_urls,
            })
    return pd.DataFrame(rows, columns=columns)

def sbom_to_excel(input_json, output_excel):
    with open(input_json, "r", encoding="utf-8") as f:
        sbom = json.load(f)

    # --- Metadata sheet ---
    metadata = sbom.get("metadata", {})
    metadata_rows = []
    metadata_rows.append(["Timestamp", metadata.get("timestamp", "Not Available")])
    metadata_rows.append(["Component Name", metadata.get("component", {}).get("name", "Not Available")])
    metadata_rows.append(["Component Version", metadata.get("component", {}).get("version", "Not Available")])
    metadata_rows.append(["Component Type", metadata.get("component", {}).get("type", "Not Available")])
    metadata_rows.append(["Component BOM Ref", metadata.get("component", {}).get("bom-ref", "Not Available")])
    tools_section = metadata.get("tools", {})
    tool_components = tools_section.get("components", [])
    tool_entries = []
    for tool in tool_components:
        if isinstance(tool, dict):
            name = tool.get("name", "Unknown Tool")
            version = tool.get("version", "")
            group = tool.get("group", "")
            author = tool.get("author", "")
            license_ids = [lic.get("license", {}).get("id") for lic in tool.get("licenses", []) if lic.get("license", {}).get("id")]
            entry_parts = [name]
            if group:
                entry_parts.append(f"(Group: {group})")
            if version:
                entry_parts.append(f"v{version}")
            if author:
                entry_parts.append(f"by {author}")
            if license_ids:
                entry_parts.append(f"[License: {', '.join(license_ids)}]")
            tool_entries.append(" ".join(entry_parts))
    if tool_entries:
        metadata_rows.append(["Tools Used", tool_entries[0]])
        for tool in tool_entries[1:]:
            metadata_rows.append(["", tool])
    else:
        metadata_rows.append(["Tools Used", "None"])
    metadata_df = pd.DataFrame(metadata_rows)
    # --- SBOM Explanation ---
    explanation_text = [
        ["SBOM and Vulnerability Report Explanation"],
        [""],
        ["SBOM Generation Process:"],
        ["This SBOM was generated from CycloneDX JSON output and converted to an Excel workbook to support review and ATO submission workflows."],
        ["The pipeline builds the image, scans dependencies and the container, merges SBOM sources, and exports this human-readable artifact."],
        [""],
        ["Workbook Structure:"],
        ["- Metadata: root component and generation metadata."],
        ["- Components: all identified software components."],
        ["- Dependencies: dependency graph edges between components."],
        ["- Vulnerabilities: security vulnerabilities detected during the scan, including severity, CVSS scores, affected packages, VEX analysis state, and advisory links."]
    ]
    df_explanation = pd.DataFrame(explanation_text)
    # --- Components sheet ---
    components = sbom.get("components", [])
    component_lookup = {}
    def flatten_components(component, parent_ref=""):
        bom_ref = component.get("bom-ref") or f"{parent_ref}|{component['name']}@{component['version']}"
        component_lookup[bom_ref] = component
        nested = component.get("components", [])
        for c in nested:
            flatten_components(c, bom_ref)
    for c in components:
        flatten_components(c)
    rows = []
    for bom_ref, component in component_lookup.items():
        name = component.get("name", "")
        version = component.get("version", "")
        group = component.get("group", "")
        full_name = f"{group + '/' if group else ''}{name}"
        desc = component.get("description", "")
        licenses = component.get("licenses", [])
        if licenses:
            license_info = ", ".join(
                [lic.get("license", {}).get("id", "") for lic in licenses if lic.get("license", {}).get("id")]
            )
            if not license_info:
                license_info = "No Assertion"
        else:
            license_info = "No Assertion"
        author = component.get("author") or "Not Listed"
        purl = component.get("purl", "")
        hash_val = ""
        for extref in component.get("externalReferences", []):
            if extref.get("type") == "distribution":
                hashes = extref.get("hashes", [])
                if hashes:
                    hash_val = f'{hashes[0]["alg"]}:{hashes[0]["content"]}'
                    break
        path_prop = next(
            (p["value"] for p in component.get("properties", []) if p["name"].endswith("package:path")), ""
        )
        rows.append({
            "Package Name": full_name,
            "Version": version,
            "Author": author,
            "License": license_info,
            "Description": desc,
            "PURL": purl,
            "Path": path_prop,
            "BOM Reference": bom_ref,
            "Hash Value": hash_val
        })
    components_df = pd.DataFrame(rows)
    if not components_df.empty:
        components_df = components_df.sort_values(by=["Package Name", "Version"])
    # --- Vulnerabilities sheet ---
    vulns_df = build_vulnerabilities_df(sbom)
    # --- Dependencies sheet ---
    dependencies = sbom.get("dependencies", [])
    dep_rows = []
    for d in dependencies:
        ref = d.get("ref", "")
        for dep in d.get("dependsOn", []):
            dep_rows.append({"Component": ref, "Depends On": dep})
    dependencies_df = pd.DataFrame(dep_rows)
    # Write all sheets using the default Excel writer (no formatting libs)
    with pd.ExcelWriter(output_excel) as writer:
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False, header=False)
        df_explanation.to_excel(writer, sheet_name="SBOM Explanation", index=False, header=False)
        components_df.to_excel(writer, sheet_name="Components", index=False)
        dependencies_df.to_excel(writer, sheet_name="Dependencies", index=False)
        vulns_df.to_excel(writer, sheet_name="Vulnerabilities", index=False)
    wb = openpyxl.load_workbook(output_excel)
    bold_font = Font(bold=True)
    ws = wb["Metadata"]
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = bold_font
    ws_vulns = wb["Vulnerabilities"]
    for col in range(1, ws_vulns.max_column + 1):
        ws_vulns.cell(row=1, column=col).font = bold_font
    wb.save(output_excel)
    print(f"Excel file created with Metadata, Components, Dependencies, and Vulnerabilities sheets: {output_excel}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert CycloneDX SBOM JSON to Excel with metadata and dependencies")
    parser.add_argument("--input", "-i", required=True, help="Path to input SBOM JSON file")
    parser.add_argument("--output", "-o", default="cyclonedx_report.xlsx", help="Path to output Excel file")
    args = parser.parse_args()
    input_file = Path(args.input)
    output_file = Path(args.output)
    if not input_file.exists():
        print(f"Input file '{input_file}' not found.")
    else:
        sbom_to_excel(input_file, output_file)
